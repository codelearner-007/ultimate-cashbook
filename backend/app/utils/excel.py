from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Palette ────────────────────────────────────────────────────────────────
TEAL        = "39AAAA"
TEAL_DARK   = "2B8080"
TEAL_LIGHT  = "F4FAFA"
TEAL_MID    = "DFF0F0"
GREEN       = "15803D"
GREEN_LIGHT = "DCFCE7"
RED         = "B91C1C"
RED_LIGHT   = "FEE2E2"
GREY        = "F8FAFC"
BORDER_CLR  = "E2E8F0"
DARK        = "0F172A"
MUTED       = "64748B"
WHITE       = "FFFFFF"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _side(style="thin", color=BORDER_CLR) -> Side:
    return Side(border_style=style, color=color)


def _border(style="thin") -> Border:
    s = _side(style)
    return Border(left=s, right=s, top=s, bottom=s)


def _thick_border() -> Border:
    th = _side("medium", TEAL)
    tn = _side("thin")
    return Border(left=tn, right=tn, top=th, bottom=th)


_PAYMENT_LABELS = {"cash": "Cash", "online": "Online", "cheque": "Cheque", "other": "Other"}
_TYPE_LABELS    = {"in": "Cash In", "out": "Cash Out"}


def _fill_row(ws, row, col_start, col_end, fill):
    for col in range(col_start, col_end + 1):
        ws.cell(row=row, column=col).fill = fill


def generate_excel(book_name: str, currency: str, entries: list, summary: dict,
                   date_from=None, date_to=None, filters: dict = None, contact_type=None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    wb.properties.creator = "Ultimate CashBook"

    sym  = (currency or "").strip()
    cfmt = f'"{sym} " #,##0.00'
    net  = summary["net_balance"]

    # ── Build filter items ────────────────────────────────────────────────────
    filter_items = []
    if date_from or date_to:
        filter_items.append(("Date Range", f"{date_from or 'Beginning'} → {date_to or 'Today'}"))
    if filters:
        if filters.get("entry_type"):
            filter_items.append(("Entry Type", _TYPE_LABELS.get(filters["entry_type"], filters["entry_type"].title())))
        if filters.get("contact_name"):
            _lbl = "Customer" if contact_type == "customer" else "Supplier" if contact_type == "supplier" else "Contact"
            filter_items.append((_lbl, filters["contact_name"]))
        if filters.get("category"):
            filter_items.append(("Category", filters["category"]))
        if filters.get("payment_mode"):
            filter_items.append(("Payment Mode", _PAYMENT_LABELS.get(filters["payment_mode"], filters["payment_mode"].title())))

    # ── Dynamic row offsets ───────────────────────────────────────────────────
    n_filter_rows  = max(1, len(filter_items))   # ≥1 (no-filter message = 1 row)
    FHDR_ROW       = 3
    FDATA_START    = 4
    FSEP_ROW       = FDATA_START + n_filter_rows  # thin separator after filters
    SUMM_LBL_ROW   = FSEP_ROW + 1
    SUMM_VAL_ROW   = SUMM_LBL_ROW + 1
    SSEP_ROW       = SUMM_VAL_ROW + 1            # thin separator after summary
    DATA_HDR_ROW   = SSEP_ROW + 1
    DATA_START_ROW = DATA_HDR_ROW + 1

    # ═══════════════════════════════════════════════════════════════════════
    # ROW 1 — App banner
    # ═══════════════════════════════════════════════════════════════════════
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = "Ultimate CashBook  —  Financial Report"
    c.font  = Font(bold=True, size=16, color=WHITE)
    c.fill  = _fill(TEAL)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30

    # ═══════════════════════════════════════════════════════════════════════
    # ROW 2 — Book & period info
    # ═══════════════════════════════════════════════════════════════════════
    ws.merge_cells("A2:I2")
    end_date = date_to or datetime.now().strftime("%Y-%m-%d")
    c = ws["A2"]
    c.value = f"  {book_name}  ·  {date_from or 'All time'}  →  {end_date}  ·  {len(entries)} transactions"
    c.font  = Font(size=9, color=WHITE)
    c.fill  = _fill(TEAL_DARK)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    # ═══════════════════════════════════════════════════════════════════════
    # ROW 3 — "APPLIED FILTERS" header
    # ═══════════════════════════════════════════════════════════════════════
    ws.merge_cells(f"A{FHDR_ROW}:I{FHDR_ROW}")
    fh = ws[f"A{FHDR_ROW}"]
    n_active   = len(filter_items)
    active_txt = f"({n_active} active)" if n_active else "None — all entries shown"
    fh.value = f"  APPLIED FILTERS  ·  {active_txt}"
    fh.font  = Font(bold=True, size=9, color=WHITE)
    fh.fill  = _fill(TEAL)
    fh.alignment = Alignment(horizontal="left", vertical="center")
    _fill_row(ws, FHDR_ROW, 2, 9, _fill(TEAL))
    ws.row_dimensions[FHDR_ROW].height = 18

    # ═══════════════════════════════════════════════════════════════════════
    # ROWS 4+ — Individual filter rows (or "no filters" message)
    # ═══════════════════════════════════════════════════════════════════════
    if filter_items:
        for i, (label, value) in enumerate(filter_items):
            row = FDATA_START + i

            ws.merge_cells(f"A{row}:C{row}")
            lc = ws.cell(row=row, column=1, value=f"  {label}")
            lc.font      = Font(bold=True, size=8.5, color=TEAL_DARK)
            lc.fill      = _fill(TEAL_MID)
            lc.alignment = Alignment(horizontal="left", vertical="center")
            _fill_row(ws, row, 2, 3, _fill(TEAL_MID))

            ws.merge_cells(f"D{row}:I{row}")
            vc = ws.cell(row=row, column=4, value=f"  {value}")
            vc.font      = Font(size=8.5, color=DARK)
            vc.fill      = _fill(TEAL_LIGHT)
            vc.alignment = Alignment(horizontal="left", vertical="center")
            _fill_row(ws, row, 5, 9, _fill(TEAL_LIGHT))

            ws.row_dimensions[row].height = 15
    else:
        row = FDATA_START
        ws.merge_cells(f"A{row}:I{row}")
        nc = ws.cell(row=row, column=1, value="  No additional filters applied — showing all entries")
        nc.font      = Font(italic=True, size=8.5, color=MUTED)
        nc.fill      = _fill(TEAL_LIGHT)
        nc.alignment = Alignment(horizontal="left", vertical="center")
        _fill_row(ws, row, 2, 9, _fill(TEAL_LIGHT))
        ws.row_dimensions[row].height = 15

    # ── Thin separator between filters and summary ────────────────────────────
    ws.merge_cells(f"A{FSEP_ROW}:I{FSEP_ROW}")
    ws.cell(row=FSEP_ROW, column=1).fill = _fill(TEAL_MID)
    _fill_row(ws, FSEP_ROW, 2, 9, _fill(TEAL_MID))
    ws.row_dimensions[FSEP_ROW].height = 4

    # ═══════════════════════════════════════════════════════════════════════
    # Summary block
    # ═══════════════════════════════════════════════════════════════════════
    summary_items = [
        ("TOTAL INCOME",   summary["total_in"],  GREEN, GREEN_LIGHT),
        ("TOTAL EXPENSES", summary["total_out"], RED,   RED_LIGHT),
        ("NET BALANCE",    net, GREEN if net >= 0 else RED,
         GREEN_LIGHT if net >= 0 else RED_LIGHT),
    ]
    for idx, (label, value, val_color, bg) in enumerate(summary_items):
        col = idx + 1

        lc = ws.cell(row=SUMM_LBL_ROW, column=col, value=label)
        lc.font      = Font(bold=True, size=8, color=WHITE)
        lc.fill      = _fill(TEAL)
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.border    = _border()

        vc = ws.cell(row=SUMM_VAL_ROW, column=col, value=value)
        vc.font          = Font(bold=True, size=13, color=val_color)
        vc.fill          = _fill(bg)
        vc.number_format = cfmt
        vc.alignment     = Alignment(horizontal="center", vertical="center")
        vc.border        = _thick_border()

    ws.row_dimensions[SUMM_LBL_ROW].height = 18
    ws.row_dimensions[SUMM_VAL_ROW].height = 24

    # ── Thin separator between summary and data headers ───────────────────────
    ws.merge_cells(f"A{SSEP_ROW}:I{SSEP_ROW}")
    ws.cell(row=SSEP_ROW, column=1).fill = _fill(TEAL_LIGHT)
    _fill_row(ws, SSEP_ROW, 2, 9, _fill(TEAL_LIGHT))
    ws.row_dimensions[SSEP_ROW].height = 6

    # ═══════════════════════════════════════════════════════════════════════
    # Column headers
    # ═══════════════════════════════════════════════════════════════════════
    HEADERS = ["Date", "Time", "Remark", "Category", "Contact", "Payment Mode",
               "Cash In", "Cash Out", "Balance"]
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=DATA_HDR_ROW, column=col, value=h)
        c.font      = Font(bold=True, color=WHITE, size=9)
        c.fill      = _fill(TEAL)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _border()
    ws.row_dimensions[DATA_HDR_ROW].height = 20

    # ═══════════════════════════════════════════════════════════════════════
    # Data rows
    # ═══════════════════════════════════════════════════════════════════════
    running   = 0.0
    total_in  = 0.0
    total_out = 0.0

    for i, e in enumerate(entries):
        row_idx = DATA_START_ROW + i
        amt     = float(e["amount"])
        is_in   = e["type"] == "in"

        if is_in:
            running  += amt; total_in += amt
            in_val, out_val = amt, None
        else:
            running  -= amt; total_out += amt
            in_val, out_val = None, amt

        row_bg = _fill(GREY) if row_idx % 2 == 0 else _fill(WHITE)
        bd     = _border("hair")

        def _d(col, val, fmt=None, bold=False, val_color=DARK, center=False,
               _r=row_idx, _bg=row_bg, _bd=bd):
            c = ws.cell(row=_r, column=col, value=val)
            c.fill      = _bg
            c.font      = Font(size=9, bold=bold, color=val_color)
            c.alignment = Alignment(horizontal="center" if center else "left",
                                    vertical="center")
            c.border    = _bd
            if fmt:
                c.number_format = fmt
            return c

        _d(1, str(e.get("entry_date", ""))[:10])
        _d(2, str(e.get("entry_time") or "")[:5], center=True)
        _d(3, e.get("remark") or "")
        _d(4, e.get("category") or "")
        _d(5, e.get("contact_name") or "")
        _d(6, e.get("payment_mode") or "")

        if in_val is not None:
            _d(7, in_val,  cfmt, bold=True, val_color=GREEN, center=True)
        else:
            _d(7, None)

        if out_val is not None:
            _d(8, out_val, cfmt, bold=True, val_color=RED, center=True)
        else:
            _d(8, None)

        bal_color = GREEN if running >= 0 else RED
        _d(9, running, cfmt, bold=True, val_color=bal_color, center=True)
        ws.row_dimensions[row_idx].height = 16

    # ═══════════════════════════════════════════════════════════════════════
    # Totals row
    # ═══════════════════════════════════════════════════════════════════════
    tr = DATA_START_ROW + len(entries)
    for col in range(1, 10):
        c = ws.cell(row=tr, column=col)
        c.fill      = _fill(TEAL)
        c.font      = Font(bold=True, color=WHITE, size=9)
        c.border    = _border()
        c.alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=tr, column=1, value="TOTAL")
    ws.cell(row=tr, column=7, value=total_in).number_format  = cfmt
    ws.cell(row=tr, column=8, value=total_out).number_format = cfmt
    ws.cell(row=tr, column=9, value=running).number_format   = cfmt
    ws.row_dimensions[tr].height = 18

    # ═══════════════════════════════════════════════════════════════════════
    # Footer note
    # ═══════════════════════════════════════════════════════════════════════
    note_row = tr + 2
    ws.merge_cells(f"A{note_row}:I{note_row}")
    nc = ws.cell(row=note_row, column=1,
                 value=f"Generated by Ultimate CashBook  ·  {datetime.now().strftime('%B %d, %Y  %I:%M %p')}")
    nc.font      = Font(size=8, color="94A3B8", italic=True)
    nc.alignment = Alignment(horizontal="left", vertical="center")

    # ═══════════════════════════════════════════════════════════════════════
    # Column widths / freeze / auto-filter
    # ═══════════════════════════════════════════════════════════════════════
    widths = [13, 8, 30, 15, 15, 16, 16, 16, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = f"A{DATA_START_ROW}"
    if entries:
        ws.auto_filter.ref = f"A{DATA_HDR_ROW}:I{tr - 1}"

    ws.sheet_properties.tabColor = TEAL

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
